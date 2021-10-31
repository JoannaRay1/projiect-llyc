# 2021.10.30
# 发现此系统存在的纰漏：
# 随机筛选中没有二次筛选年级

import openpyxl

# 将数据中的年级转化为数字，方便评分时进行比较
#修改处1：更换了年级数字化的方式
def grade_digitalize(ws_value):
    for each in ws_value:
         tmp=list()
    for each in ws_value:
        # 本人年级数字化
        # 期望对方年级数字化且范围化  假设each[6]为最高年级，each[5]为最低年级        
        grade={'大一':1,'大二':2,'大三':3,'大四':4,'硕士':5}
        tmp.append([grade[i] if i in grade else i for i in each])
    ws_value=tmp          
    return ws_value


# 按照性取向分组
# 男异性恋
def boy_heterosexual(line):
    if line[2] != line[3] and line[2] == '男生':
        return True
    return False


# 男同性恋
def boy_homosexual(line):
    if line[2] == line[3] and line[2] == '男生':
        return True
    return False


# 女异性恋
def girl_heterosexual(line):
    if line[2] != line[3] and line[2] == '女生':
        return True
    return False


# 女同性恋
def girl_homosexual(line):
    if line[2] == line[3] and line[2] == '女生':
        return True
    return False


# 按性取向分组并返回
def gender_orientation_initialize(ws_value):
    # 此处可用if嵌套进行优化 没那么多冗余 但现在这样代码容易写 我懒:)
    boy_straight = list(filter(boy_heterosexual, ws_value))
    boy_gay = list(filter(boy_homosexual, ws_value))
    girl_straight = list(filter(girl_heterosexual, ws_value))
    girl_gay = list(filter(girl_homosexual, ws_value))
    return boy_straight, boy_gay, girl_straight, girl_gay

#修改处2：增加grade_match函数
#判断年级是否完全匹配成功
def grade_match(person1,person2):
    if (person1[6] >= person2[4]) and (person1[5] <= person2[4] ):
        if (person2[6] >= person1[4] ) and person2[5] <= (person1[4] ):
            return True
    return False        
        
        
# 匹配程度打分
def match_degree(person1, person2,n):
    result = 0
    i=1;
#修改处3：若年级不匹配得分直接为0
    if not(grade_match(person1, person2)):
        return 0    
    # 按照条件进行双向打分
      
    while(i<=n):
        if person1[5+i] == person2[6+i]:
            result += 1
        ++i    
    i=1
    while(i<=n):
        if person1[5+i] == person2[6+i]:
            result += 1
        ++i
    return result


# 异性恋条件匹配
#修改处4：引入匹配条件个数n为新参数
#修改处5：current_degree的两个临界值取与n有关参数
def condition_match(group1, group2, num, group_num, final_sheet,n):
    remain_group1 = list()
    to_match = dict()
    
    for person in group1:  # 正向筛选，选择对person1来说高分的person2
        current_degree = 0  # 匹配程度初始化
        current_person = []  # 匹配对象初始化
        if len(group2) < 1:
            break
        for person2 in group2:
            former_degree = current_degree
            current_degree = max(former_degree, match_degree(person, person2,n))  # 取高分对象
            if current_degree == 2*n+2 :  # 得到满分直接匹配
                current_person = person2
                break
            if current_degree > former_degree:
                current_person = person2  # 取高分对象
        if current_degree > 2*n-1 :
            # 此处取2n-1的原因：
            # ① 重视年级所占分数。
            # ② 如果两个人的年级没有任何一方是匹配的，除非其他所有条件全部互选，否则不可能在这里被选择。
            final_sheet.cell(num, 1, person[0])
            final_sheet.cell(num, 2, person[1])
            final_sheet.cell(num, 3, group_num)
            final_sheet.cell(num, 4, current_person[1])
            num += 1
            final_sheet.cell(num, 1, current_person[0])
            final_sheet.cell(num, 2, current_person[1])
            final_sheet.cell(num, 3, group_num)
            final_sheet.cell(num, 4, person[1])
            num += 1
            group_num += 1
            group2.remove(current_person)
        else:
            remain_group1.append(person)
            if to_match.get(current_person[0]):  # 反向筛选，存入对person2来说最高分的person1    
                if to_match[current_person[0]][1] < current_degree:
                    to_match[current_person[0]] = [person, current_degree]
            else:
                to_match[current_person[0]] = [person, current_degree]
    for key in to_match:  # 反向筛选结果进入最终list
        t_p = []
        for p in group2:
            if key in p:
                t_p = p
                break
        if len(t_p) < 1:
            continue
        final_sheet.cell(num, 1, to_match[key][0][0])
        final_sheet.cell(num, 2, to_match[key][0][1])
        final_sheet.cell(num, 3, group_num)
        final_sheet.cell(num, 4, t_p[1])
        num += 1
        final_sheet.cell(num, 1, t_p[0])
        final_sheet.cell(num, 2, t_p[1])
        final_sheet.cell(num, 3, group_num)
        final_sheet.cell(num, 4, to_match[key][0][1])
        num += 1
        group_num += 1
        group2.remove(t_p)
        remain_group1.remove(to_match[key][0])
    return num, group_num, remain_group1, group2


    
# 异性恋匹配分数不足的进行随机匹配
def random_match(group1, group2, num, group_num, final_sheet, no_match):
    length = min(len(group1), len(group2))
    while length > 0:
        final_sheet.cell(num, 1, group1[0][0])
        final_sheet.cell(num, 2, group1[0][1])
        final_sheet.cell(num, 3, group_num)
        final_sheet.cell(num, 4, group2[0][1])
        num += 1
        final_sheet.cell(num, 1, group2[0][0])
        final_sheet.cell(num, 2, group2[0][1])
        final_sheet.cell(num, 3, group_num)
        final_sheet.cell(num, 4, group1[0][1])
        num += 1
        group_num += 1
        length -= 1
        group1.pop(0)
        group2.pop(0)
    while len(group1) > 0:
        t_p = group1.pop(0)
        no_match.append(t_p)
    while len(group2) > 0:
        t_p = group2.pop(0)
        no_match.append(t_p)
    return num, group_num, no_match


# 同性恋条件+随机匹配
# 考虑到每次参加的LGBT人群数量较少，因此没有进行循环查找匹配
def lgbt_match(group1, group2, num, group_num, final_sheet, no_match):
    remain_group = list()
    while len(group1) > 0:
        t_p = group1.pop(0)
        current_degree = 0
        t_tp = []
        for person in group2:
            former_degree = current_degree
            current_degree = max(current_degree, match_degree(t_p, person,n))
            if current_degree > former_degree:
                t_tp = person
        if current_degree > 5:
            final_sheet.cell(num, 1, t_p[0])
            final_sheet.cell(num, 2, t_p[1])
            final_sheet.cell(num, 3, group_num)
            final_sheet.cell(num, 4, t_tp[1])
            num += 1
            final_sheet.cell(num, 1, t_tp[0])
            final_sheet.cell(num, 2, t_tp[1])
            final_sheet.cell(num, 3, group_num)
            final_sheet.cell(num, 4, t_p[1])
            num += 1
            group_num += 1
            group1.remove(t_tp)
        else:
            remain_group.append(t_p)
    while len(remain_group) > 1:
        final_sheet.cell(num, 1, remain_group[0][0])
        final_sheet.cell(num, 2, remain_group[0][1])
        final_sheet.cell(num, 3, group_num)
        final_sheet.cell(num, 4, remain_group[1][1])
        num += 1
        final_sheet.cell(num, 1, remain_group[1][0])
        final_sheet.cell(num, 2, remain_group[1][1])
        final_sheet.cell(num, 3, group_num)
        final_sheet.cell(num, 4, remain_group[0][1])
        num += 1
        group_num += 1
        remain_group.pop(0)
        remain_group.pop(0)
    no_match.extend(remain_group)
    return num, group_num, no_match


# 剩余的没有匹配的人员放入表中
def finish(group, num, final_sheet):
    while len(group) > 0:
        final_sheet.cell(num, 1, group[0][0])
        final_sheet.cell(num, 2, group[0][1])
        final_sheet.cell(num, 3, '无')
        final_sheet.cell(num, 4, '抱歉，您本次没有获得匹配。')
        num += 1
        group.pop(0)
    return


if __name__ == '__main__':
    # path = input('请输入需要进行匹配的文件地址：(注意后缀名)')
    path = '恋爱觉醒.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    # 获得ws数值列表
    ws_value = list()
    for each in ws:
        line_value = list()
        for each_cell in each:
            line_value.append(each_cell.value)
        ws_value.append(line_value)
    ws_value = grade_digitalize(ws_value)
    boy_straight, boy_gay, girl_straight, girl_gay = gender_orientation_initialize(ws_value)
    # print([boy_straight, boy_gay, girl_straight, girl_gay])
    #对于match_degree()的条件个数由n做决定   
    n=int (input( "请输入除性别年级外要匹配的条件个数:"))
    final_sheet = wb.create_sheet('Sheet2', 1)
    num = 1
    group_num = 1
    no_match = list()
    row_num, existing_group, remain_girlS, remain_boyS = condition_match(girl_straight, boy_straight, num, group_num, final_sheet,n)
    row_num, existing_group, no_match = random_match(remain_girlS, remain_boyS, row_num, existing_group, final_sheet, no_match)
    row_num, existing_group, no_match = lgbt_match(boy_gay, boy_gay, row_num, existing_group, final_sheet, no_match)
    row_num, existing_group, no_match = lgbt_match(girl_gay, girl_gay, row_num, existing_group, final_sheet, no_match)
    finish(no_match, row_num, final_sheet)
    wb.save(path)
